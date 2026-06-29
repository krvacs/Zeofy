import os
import random
import warnings
import joblib

from copy import deepcopy

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.model_selection import (
    train_test_split,
    StratifiedKFold
)

from sklearn.preprocessing import (
    StandardScaler,
    LabelEncoder
)

from sklearn.metrics import (
    precision_score,
    recall_score,
    f1_score,
    accuracy_score,
    confusion_matrix,
    ConfusionMatrixDisplay
)

from sklearn.inspection import (
    permutation_importance
)

# NEW IMPORT: Required for newer Scikit-Learn versions
from sklearn.base import (
    BaseEstimator,
    ClassifierMixin
)

# ─────────────────────────────────────────
# PYTORCH
# ─────────────────────────────────────────
import torch
import torch.nn as nn
import torch.optim as optim

from torch.utils.data import (
    TensorDataset,
    DataLoader
)

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────
# RANDOM SEED
# ─────────────────────────────────────────
GLOBAL_SEED = 42

random.seed(GLOBAL_SEED)
np.random.seed(GLOBAL_SEED)
torch.manual_seed(GLOBAL_SEED)

if torch.cuda.is_available():
    torch.cuda.manual_seed_all(GLOBAL_SEED)

DEVICE = torch.device(
    "cuda" if torch.cuda.is_available() else "cpu"
)

print(f"\nUsing device: {DEVICE}")

# ─────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────
df = pd.read_csv("versions/v2/datasets/2nd - Experiment/draft_zeolite6.csv")

# ─────────────────────────────────────────
# FEATURE ENGINEERING
# ─────────────────────────────────────────
df['si_al_ratio'] = np.log1p(
    df['sival'] / (df['alval'] + 1e-6)
)

df['oh_si_ratio'] = np.log1p(
    df['ohval'] / (df['sival'] + 1e-6)
)

df['na_al_ratio'] = np.log1p(
    df['naval'] / (df['alval'] + 1e-6)
)

df['oh_al_ratio'] = np.log1p(
    df['ohval'] / (df['alval'] + 1e-6)
)

feature_names = [
    'sival',
    'alval',
    'naval',
    'h20',
    'mag',
    'ohval',
    'time',
    'temper',
    'si_al_ratio',
    'oh_si_ratio',
    'na_al_ratio',
    'oh_al_ratio'
]

X = df[feature_names]
y = df['code1']

# ─────────────────────────────────────────
# LABEL ENCODING
# ─────────────────────────────────────────
le = LabelEncoder()

y = le.fit_transform(y)

num_classes = len(np.unique(y))
input_size = X.shape[1]

# ─────────────────────────────────────────
# TRAIN TEST SPLIT
# ─────────────────────────────────────────
X_trainval, X_test, y_trainval, y_test = train_test_split(
    X,
    y,
    test_size=0.15,
    stratify=y,
    random_state=42
)

# ─────────────────────────────────────────
# STANDARDIZATION
# ─────────────────────────────────────────
scaler = StandardScaler()

X_trainval = scaler.fit_transform(
    X_trainval
)

X_test = scaler.transform(
    X_test
)

# ─────────────────────────────────────────
# HYPERPARAMETER SPACE
# ─────────────────────────────────────────
HYPERPARAMETER_SPACE = {
    'n_hidden_layers': [1, 2, 3],
    'hidden_size_1': [32, 64, 128],
    'hidden_size_2': [16, 32, 64],
    'hidden_size_3': [16, 32],
    'hidden_activation': ['relu', 'elu', 'tanh'],
    'optimizer_type': ['AdamW', 'RMSprop', 'SGD'],
    'learning_rate': [0.0001, 0.0005, 0.001],
    'weight_decay': [0.0001, 0.001, 0.01],
    'momentum': [0.9, 0.95, 0.99],
    'dropout_rate': [0.0, 0.1, 0.2, 0.3],
    'use_batch_norm': [True],
    'early_stop_patience': [20, 30, 50],
    'batch_size': [16, 32, 64],
    'epochs': [300, 500],
    'lr_scheduler': ['plateau', 'cosine', 'none'],
    'plateau_patience': [5, 10, 15],
    'step_size': [30, 50, 75],
    'gamma': [0.1, 0.5, 0.7],
    'weight_init': ['xavier_uniform', 'he_uniform'],
    'bias_init': ['zeros', 'constant'],
    'random_seed': [42]
}

# ─────────────────────────────────────────
# ACTIVATION FUNCTION
# ─────────────────────────────────────────
def get_activation(name):
    if name == "relu":
        return nn.ReLU()
    elif name == "elu":
        return nn.ELU()
    elif name == "tanh":
        return nn.Tanh()
    return nn.ReLU()

# ─────────────────────────────────────────
# WEIGHT INITIALIZATION
# ─────────────────────────────────────────
def initialize_weights(model, weight_init, bias_init):
    for m in model.modules():
        if isinstance(m, nn.Linear):
            if weight_init == "xavier_uniform":
                nn.init.xavier_uniform_(m.weight)
            elif weight_init == "he_uniform":
                nn.init.kaiming_uniform_(m.weight, nonlinearity='relu')
            
            if bias_init == "zeros":
                nn.init.zeros_(m.bias)
            elif bias_init == "constant":
                nn.init.constant_(m.bias, 0.01)

# ─────────────────────────────────────────
# NEURAL NETWORK
# ─────────────────────────────────────────
class ZeoliteNN(nn.Module):
    def __init__(self, input_size, num_classes, params):
        super(ZeoliteNN, self).__init__()
        layers = []
        activation = get_activation(params['hidden_activation'])
        hidden_sizes = []
        hidden_sizes.append(params['hidden_size_1'])
        
        if params['n_hidden_layers'] >= 2:
            hidden_sizes.append(params['hidden_size_2'])
        if params['n_hidden_layers'] >= 3:
            hidden_sizes.append(params['hidden_size_3'])

        prev_size = input_size
        for hidden_size in hidden_sizes:
            layers.append(nn.Linear(prev_size, hidden_size))
            if params['use_batch_norm']:
                layers.append(nn.BatchNorm1d(hidden_size))
            layers.append(activation)
            if params['dropout_rate'] > 0:
                layers.append(nn.Dropout(params['dropout_rate']))
            prev_size = hidden_size

        layers.append(nn.Linear(prev_size, num_classes))
        self.network = nn.Sequential(*layers)
        initialize_weights(self.network, params['weight_init'], params['bias_init'])

    def forward(self, x):
        return self.network(x)

# ─────────────────────────────────────────
# SCIKIT-LEARN PYTORCH WRAPPER
# ─────────────────────────────────────────
class PyTorchSklearnWrapper(BaseEstimator, ClassifierMixin):
    """Tricks Scikit-Learn into thinking the PyTorch model is a standard SKLearn classifier."""
    
    def __init__(self, model=None, device='cpu', classes_=None):
        self.model = model
        self.device = device
        self.classes_ = classes_
        
    def fit(self, X, y):
        # We don't need to fit it again here, just return self
        return self
        
    def predict(self, X):
        self.model.eval()
        with torch.no_grad():
            if not isinstance(X, torch.Tensor):
                X = torch.tensor(X, dtype=torch.float32)
            X = X.to(self.device)
            outputs = self.model(X)
            return torch.argmax(outputs, dim=1).cpu().numpy()

# ─────────────────────────────────────────
# METRIC FUNCTION
# ─────────────────────────────────────────
def calculate_metrics(y_true, y_pred):
    return {
        'accuracy': accuracy_score(y_true, y_pred),
        'precision': precision_score(y_true, y_pred, average='weighted', zero_division=0),
        'recall': recall_score(y_true, y_pred, average='weighted', zero_division=0),
        'f1_score': f1_score(y_true, y_pred, average='weighted', zero_division=0)
    }

# ─────────────────────────────────────────
# TRAIN FUNCTION
# ─────────────────────────────────────────
def train_model(model, train_loader, val_loader, params):
    criterion = nn.CrossEntropyLoss()
    train_losses = []
    val_losses = []

    # OPTIMIZER
    if params['optimizer_type'] == 'AdamW':
        optimizer = optim.AdamW(model.parameters(), lr=params['learning_rate'], weight_decay=params['weight_decay'])
    elif params['optimizer_type'] == 'RMSprop':
        optimizer = optim.RMSprop(model.parameters(), lr=params['learning_rate'], weight_decay=params['weight_decay'], momentum=params['momentum'])
    elif params['optimizer_type'] == 'SGD':
        optimizer = optim.SGD(model.parameters(), lr=params['learning_rate'], momentum=params['momentum'], weight_decay=params['weight_decay'])

    scheduler = None
    if params['lr_scheduler'] == 'plateau':
        scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', patience=params['plateau_patience'])
    elif params['lr_scheduler'] == 'step':
        scheduler = optim.lr_scheduler.StepLR(optimizer, step_size=params['step_size'], gamma=params['gamma'])
    elif params['lr_scheduler'] == 'cosine':
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=params['epochs'])

    best_val_loss = np.inf
    best_model_state = None
    patience_counter = 0

    for epoch in range(params['epochs']):
        # TRAIN
        model.train()
        train_loss = 0.0
        for X_batch, y_batch in train_loader:
            X_batch, y_batch = X_batch.to(DEVICE), y_batch.to(DEVICE)
            optimizer.zero_grad()
            outputs = model(X_batch)
            loss = criterion(outputs, y_batch)
            loss.backward()
            nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()
            train_loss += loss.item()
        
        train_loss /= max(len(train_loader), 1)

        # VALIDATION
        model.eval()
        val_loss = 0.0
        with torch.no_grad():
            for X_batch, y_batch in val_loader:
                X_batch, y_batch = X_batch.to(DEVICE), y_batch.to(DEVICE)
                outputs = model(X_batch)
                loss = criterion(outputs, y_batch)
                val_loss += loss.item()
        
        val_loss /= max(len(val_loader), 1)
        train_losses.append(train_loss)
        val_losses.append(val_loss)

        if scheduler:
            if params['lr_scheduler'] == 'plateau':
                scheduler.step(val_loss)
            else:
                scheduler.step()

        if val_loss < best_val_loss:
            best_val_loss = val_loss
            best_model_state = deepcopy(model.state_dict())
            patience_counter = 0
        else:
            patience_counter += 1

        if patience_counter >= params['early_stop_patience']:
            break

    if best_model_state:
        model.load_state_dict(best_model_state)

    return model, train_losses, val_losses

def sample_hyperparameters(space):
    sampled = {}
    for key, values in space.items():
        sampled[key] = random.choice(values)
    return sampled

# ─────────────────────────────────────────
# EXCEL SETUP 
# ─────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = "saved_model/results"
os.makedirs(RESULTS_DIR, exist_ok=True)
excel_path = os.path.join(RESULTS_DIR, "best_model_results.xlsx")

hp_cols    = [f"hp_{k}" for k in HYPERPARAMETER_SPACE.keys()]
train_cols = ['train_accuracy', 'train_precision', 'train_recall', 'train_f1_score']

# UPDATED: Added mean and std columns for CV metrics
cv_cols    = [
    'cv_accuracy_mean', 'cv_accuracy_std', 
    'cv_precision_mean', 'cv_precision_std', 
    'cv_recall_mean', 'cv_recall_std', 
    'cv_f1_score_mean', 'cv_f1_score_std'
]

test_cols  = ['test_accuracy', 'test_precision', 'test_recall', 'test_f1_score']
ordered_cols = (['config_id'] + hp_cols + train_cols + cv_cols + test_cols + ['is_best'])

# UPDATED: Added display names for the new CV columns
DISPLAY_NAMES = {
    'config_id': 'Config ID', 'is_best': 'Is Best',
    'train_accuracy': 'Train Accuracy', 'train_precision': 'Train Precision', 'train_recall': 'Train Recall', 'train_f1_score': 'Train F1 Score',
    'cv_accuracy_mean': 'CV Acc (Mean)', 'cv_accuracy_std': 'CV Acc (Std)', 
    'cv_precision_mean': 'CV Prec (Mean)', 'cv_precision_std': 'CV Prec (Std)', 
    'cv_recall_mean': 'CV Recall (Mean)', 'cv_recall_std': 'CV Recall (Std)', 
    'cv_f1_score_mean': 'CV F1 (Mean)', 'cv_f1_score_std': 'CV F1 (Std)',
    'test_accuracy': 'Test Accuracy', 'test_precision': 'Test Precision', 'test_recall': 'Test Recall', 'test_f1_score': 'Test F1 Score',
}
for k in HYPERPARAMETER_SPACE.keys():
    DISPLAY_NAMES[f"hp_{k}"] = k

HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HP_FILL    = PatternFill('solid', fgColor='D6E4F0')
TRAIN_FILL = PatternFill('solid', fgColor='E2EFDA')
CV_FILL    = PatternFill('solid', fgColor='FFF2CC')
TEST_FILL  = PatternFill('solid', fgColor='FCE4D6')
BEST_FILL  = PatternFill('solid', fgColor='C6EFCE')

WHITE_BOLD   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BOLD_FONT    = Font(name='Arial', bold=True, size=10)
NORMAL_FONT  = Font(name='Arial', size=10)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
_side        = Side(style='thin', color='BFBFBF')
BORDER       = Border(left=_side, right=_side, top=_side, bottom=_side)

HP_END    = 1 + len(hp_cols)
TRAIN_END = HP_END + len(train_cols)
CV_END    = TRAIN_END + len(cv_cols)
TEST_END  = CV_END + len(test_cols)

def _col_fill(col_idx):
    if col_idx == 1: return None
    elif col_idx <= HP_END: return HP_FILL
    elif col_idx <= TRAIN_END: return TRAIN_FILL
    elif col_idx <= CV_END: return CV_FILL
    elif col_idx <= TEST_END: return TEST_FILL
    return None

def _init_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Random Search Trials'
    for col_idx, col_key in enumerate(ordered_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=DISPLAY_NAMES.get(col_key, col_key))
        cell.font, cell.fill, cell.alignment, cell.border = WHITE_BOLD, HDR_FILL, CENTER, BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(DISPLAY_NAMES.get(col_key, col_key)) + 4, 14)
    ws.freeze_panes = 'B2'
    wb.create_sheet('Feature Importance')
    wb.save(path)

def _append_trial_row(path, trial_row, is_best):
    wb = openpyxl.load_workbook(path)
    ws = wb['Random Search Trials']
    next_row = ws.max_row + 1
    for col_idx, col_key in enumerate(ordered_cols, start=1):
        val = trial_row.get(col_key, '')
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.alignment, cell.border = CENTER, BORDER
        if is_best:
            cell.fill, cell.font = BEST_FILL, BOLD_FONT
        else:
            fill = _col_fill(col_idx)
            if fill: cell.fill = fill
            cell.font = BOLD_FONT if col_idx == 1 else NORMAL_FONT
        col_letter = get_column_letter(col_idx)
        current_w = ws.column_dimensions[col_letter].width or 14
        needed_w = min(len(str(val or '')) + 4, 30)
        if needed_w > current_w:
            ws.column_dimensions[col_letter].width = needed_w
    wb.save(path)

X_test_tensor = torch.tensor(X_test, dtype=torch.float32)
_init_excel(excel_path)

# ─────────────────────────────────────────
# RANDOM SEARCH
# ─────────────────────────────────────────
N_RANDOM_SEARCH = 463  # Reduced for standard testing, adjust back to 400 as needed
print("\nRunning Random Search...")
best_f1 = -1
best_params = None
best_cv_metrics = None
all_trials_log = []

skf_search = StratifiedKFold(n_splits=10, shuffle=True, random_state=42)

for trial in range(N_RANDOM_SEARCH):
    print(f"\nTRIAL {trial+1}/{N_RANDOM_SEARCH}")
    params = sample_hyperparameters(HYPERPARAMETER_SPACE)
    fold_accuracy, fold_precision, fold_recall, fold_f1 = [], [], [], []

    for fold, (train_idx, val_idx) in enumerate(skf_search.split(X_trainval, y_trainval)):
        X_train, X_val = X_trainval[train_idx], X_trainval[val_idx]
        y_train, y_val = y_trainval[train_idx], y_trainval[val_idx]

        train_loader = DataLoader(
            TensorDataset(torch.tensor(X_train, dtype=torch.float32), torch.tensor(y_train, dtype=torch.long)),
            batch_size=params['batch_size'], shuffle=True, drop_last=len(X_train) > params['batch_size']
        )
        val_loader = DataLoader(
            TensorDataset(torch.tensor(X_val, dtype=torch.float32), torch.tensor(y_val, dtype=torch.long)),
            batch_size=params['batch_size'], shuffle=False
        )

        model = ZeoliteNN(input_size=input_size, num_classes=num_classes, params=params).to(DEVICE)
        model, _, _ = train_model(model, train_loader, val_loader, params)
        model.eval()

        with torch.no_grad():
            outputs = model(torch.tensor(X_val, dtype=torch.float32).to(DEVICE))
            preds = torch.argmax(outputs, dim=1).cpu().numpy()

        metrics = calculate_metrics(y_val, preds)
        fold_accuracy.append(metrics['accuracy'])
        fold_precision.append(metrics['precision'])
        fold_recall.append(metrics['recall'])
        fold_f1.append(metrics['f1_score'])

    # UPDATED: Calculate both means and stds for CV metrics
    avg_accuracy = np.mean(fold_accuracy)
    std_accuracy = np.std(fold_accuracy)
    
    avg_precision = np.mean(fold_precision)
    std_precision = np.std(fold_precision)
    
    avg_recall = np.mean(fold_recall)
    std_recall = np.std(fold_recall)
    
    avg_f1 = np.mean(fold_f1)
    std_f1 = np.std(fold_f1)
    
    print(f"Average F1: {avg_f1:.4f} ± {std_f1:.4f}")

    print(f"  Evaluating on test set...")
    trial_X_tr, trial_X_val, trial_y_tr, trial_y_val = train_test_split(X_trainval, y_trainval, test_size=0.1, stratify=y_trainval, random_state=42)
    
    trial_train_loader = DataLoader(
        TensorDataset(torch.tensor(trial_X_tr, dtype=torch.float32), torch.tensor(trial_y_tr, dtype=torch.long)),
        batch_size=params['batch_size'], shuffle=True, drop_last=len(trial_X_tr) > params['batch_size']
    )
    trial_val_loader = DataLoader(
        TensorDataset(torch.tensor(trial_X_val, dtype=torch.float32), torch.tensor(trial_y_val, dtype=torch.long)),
        batch_size=params['batch_size'], shuffle=False
    )

    trial_model = ZeoliteNN(input_size=input_size, num_classes=num_classes, params=params).to(DEVICE)
    trial_model, trial_train_losses, trial_val_losses = train_model(trial_model, trial_train_loader, trial_val_loader, params)
    trial_model.eval()

    with torch.no_grad():
        trial_test_preds = torch.argmax(trial_model(X_test_tensor.to(DEVICE)), dim=1).cpu().numpy()
    trial_test_metrics = calculate_metrics(y_test, trial_test_preds)

    with torch.no_grad():
        trial_train_preds = torch.argmax(trial_model(torch.tensor(X_trainval, dtype=torch.float32).to(DEVICE)), dim=1).cpu().numpy()
    trial_train_metrics = calculate_metrics(y_trainval, trial_train_preds)

    print(f"  Train F1: {trial_train_metrics['f1_score']:.4f}  Acc: {trial_train_metrics['accuracy']:.4f}  |  Test F1: {trial_test_metrics['f1_score']:.4f}  Acc: {trial_test_metrics['accuracy']:.4f}")

    config_id = f"Config_{trial + 1:03d}"
    is_best = avg_f1 > best_f1
    
    # ─────────────────────────────────────────
    # ARTIFACT GENERATION FOR BEST MODEL
    # ─────────────────────────────────────────
    if is_best:
        best_f1 = avg_f1
        best_params = params
        best_config_id = config_id
        print(f"NEW BEST MODEL FOUND: {config_id}")
        
        # Create a specific directory for this config
        config_dir = os.path.join("saved_model", config_id)
        os.makedirs(config_dir, exist_ok=True)
        
        # Save Model Artifacts
        torch.save(trial_model.state_dict(), os.path.join(config_dir, "neural_network_model.pth"))
        joblib.dump(scaler, os.path.join(config_dir, "scaler.pkl"))
        joblib.dump(le, os.path.join(config_dir, "label_encoder.pkl"))
        joblib.dump(params, os.path.join(config_dir, "best_params.pkl"))
        
        # 1. Learning Curve
        plt.figure(figsize=(10, 6))
        plt.plot(trial_train_losses, label='Training Loss', color='#1F4E79', linewidth=2)
        if trial_val_losses:
            plt.plot(trial_val_losses, label='Validation Loss', color='#D6E4F0', linewidth=2)
        plt.title(f'{config_id} Learning Curve', fontsize=14, fontweight='bold')
        plt.xlabel('Epochs', fontsize=12)
        plt.ylabel('Loss', fontsize=12)
        plt.legend()
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(os.path.join(config_dir, f"{config_id}_learning_curve.png"), dpi=300)
        plt.close()

        # 2. Confusion Matrix
        cm = confusion_matrix(y_test, trial_test_preds)
        disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=le.classes_)
        fig, ax = plt.subplots(figsize=(12, 10))
        disp.plot(cmap='Blues', ax=ax, xticks_rotation=45)
        plt.title(f'{config_id} Confusion Matrix (Test Set)', fontsize=14, fontweight='bold')
        plt.tight_layout()
        plt.savefig(os.path.join(config_dir, f"{config_id}_confusion_matrix.png"), dpi=300)
        plt.close()

        # 3. Feature Importance
        print("  - Calculating Feature Importance for new best model...")
        
        wrapped_model = PyTorchSklearnWrapper(model=trial_model, device=DEVICE, classes_=le.classes_)
        
        perm_importance = permutation_importance(
            wrapped_model, X_test, y_test, scoring='f1_weighted', n_repeats=10, random_state=42, n_jobs=1
        )
        sorted_idx = perm_importance.importances_mean.argsort()

        plt.figure(figsize=(10, 8))
        plt.barh(np.array(feature_names)[sorted_idx], perm_importance.importances_mean[sorted_idx], color='#1F4E79')
        plt.xlabel("Permutation Importance (Weighted F1-Score Decrease)", fontsize=12)
        plt.title(f"Feature Importance for {config_id}", fontsize=14, fontweight='bold')
        plt.grid(axis='x', linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(os.path.join(config_dir, f"{config_id}_feature_importance.png"), dpi=300)
        plt.close()

    trial_row = {'config_id': config_id, 'is_best': is_best}
    for k, v in params.items(): trial_row[f"hp_{k}"] = v
    
    trial_row['train_accuracy'] = trial_train_metrics['accuracy']
    trial_row['train_precision'] = trial_train_metrics['precision']
    trial_row['train_recall'] = trial_train_metrics['recall']
    trial_row['train_f1_score'] = trial_train_metrics['f1_score']
    
    # UPDATED: Mapping means and stds to the trial row
    trial_row['cv_accuracy_mean'] = avg_accuracy
    trial_row['cv_accuracy_std'] = std_accuracy
    
    trial_row['cv_precision_mean'] = avg_precision
    trial_row['cv_precision_std'] = std_precision
    
    trial_row['cv_recall_mean'] = avg_recall
    trial_row['cv_recall_std'] = std_recall
    
    trial_row['cv_f1_score_mean'] = avg_f1
    trial_row['cv_f1_score_std'] = std_f1
    
    trial_row['test_accuracy'] = trial_test_metrics['accuracy']
    trial_row['test_precision'] = trial_test_metrics['precision']
    trial_row['test_recall'] = trial_test_metrics['recall']
    trial_row['test_f1_score'] = trial_test_metrics['f1_score']

    all_trials_log.append(trial_row)
    _append_trial_row(excel_path, trial_row, is_best)

print("\n" + "=" * 60)
print(f"SEARCH COMPLETE. BEST MODEL: {best_config_id}")
print("=" * 60)
for k, v in best_params.items():
    print(f"{k:<25}: {v}")
print(f"\nAll artifacts for the best model have been saved in: saved_model/{best_config_id}/")

# ─────────────────────────────────────────
# BULK PREDICTION SCRIPT
# ─────────────────────────────────────────
def bulk_predict_excel(file_path):
    print("\n" + "=" * 50)
    print(f"RUNNING BULK PREDICTION ON: {file_path}")
    print("=" * 50)
    
    if not os.path.exists(file_path):
        print(f"File {file_path} not found. Skipping bulk prediction.")
        return

    # Load the absolute best model from the search for prediction
    final_model_path = os.path.join("saved_model", best_config_id, "neural_network_model.pth")
    pred_model = ZeoliteNN(input_size=input_size, num_classes=num_classes, params=best_params).to(DEVICE)
    pred_model.load_state_dict(torch.load(final_model_path))
    pred_model.eval()

    # 1. Load Data
    bulk_df = pd.read_excel(file_path)

    # 2. Recreate exact feature engineering (Log Transformed!)
    bulk_df['si_al_ratio'] = np.log1p(bulk_df['sival'] / (bulk_df['alval'] + 1e-6))
    bulk_df['oh_si_ratio'] = np.log1p(bulk_df['ohval'] / (bulk_df['sival'] + 1e-6))
    bulk_df['na_al_ratio'] = np.log1p(bulk_df['naval'] / (bulk_df['alval'] + 1e-6))
    bulk_df['oh_al_ratio'] = np.log1p(bulk_df['ohval'] / (bulk_df['alval'] + 1e-6))

    # 3. Extract exact features
    X_bulk = bulk_df[feature_names].values

    # 4. Scale EXACTLY like training
    X_bulk_scaled = scaler.transform(X_bulk)
    X_bulk_tensor = torch.tensor(X_bulk_scaled, dtype=torch.float32).to(DEVICE)

    # 5. Predict using evaluation mode
    with torch.no_grad():
        outputs = pred_model(X_bulk_tensor)
        probabilities = torch.softmax(outputs, dim=1)
        preds = torch.argmax(probabilities, dim=1).cpu().numpy()
        confidences = torch.max(probabilities, dim=1).values.cpu().numpy() * 100

    # 6. Save results
    bulk_df['Predicted Framework'] = le.inverse_transform(preds)
    bulk_df['Confidence Score'] = confidences

    out_path = file_path.replace(".xlsx", "_predictions.xlsx")
    bulk_df.to_excel(out_path, index=False)
    print(f"Bulk predictions saved successfully to: {out_path}")

# Run the bulk prediction (change file path as needed)
# bulk_predict_excel("versions/v2/datasets/zeosyn5.xlsx")


# ─────────────────────────────────────────
# USER INPUT PREDICTION 
# ─────────────────────────────────────────
print("\n" + "=" * 50)
print("ZEOLITE PHASE PREDICTOR")
print("=" * 50)

# Load the absolute best model from the search for prediction
final_model_path = os.path.join("saved_model", best_config_id, "neural_network_model.pth")
pred_model = ZeoliteNN(input_size=input_size, num_classes=num_classes, params=best_params).to(DEVICE)
pred_model.load_state_dict(torch.load(final_model_path))
pred_model.eval()

while True:
    print("\nEnter raw composition values (or type 'exit'):\n")
    try:
        sival_in = input("SiO2 (sival): ")
        if sival_in.lower() == 'exit': break
        
        sival = float(sival_in)
        alval = float(input("Al2O3 (alval): "))
        naval = float(input("Na2O (naval): "))
        mag = float(input("MgO (mag): "))
        h20 = float(input("H2O (h20): "))
        ohval = float(input("OH- (oh): "))
        time = float(input("Time (hrs): "))
        temper = float(input("Temp (C): "))
    except ValueError:
        print("Invalid numeric input.")
        continue

    # Apply log transformation to user inputs
    si_al_ratio = np.log1p(sival / (alval + 1e-6))
    oh_si_ratio = np.log1p(ohval / (sival + 1e-6))
    na_al_ratio = np.log1p(naval / (alval + 1e-6))
    oh_al_ratio = np.log1p(ohval / (alval + 1e-6))

    X_input = np.array([[
        sival, alval, naval, mag, h20, ohval, 
        time, temper, si_al_ratio, oh_si_ratio, 
        na_al_ratio, oh_al_ratio
    ]])

    # Apply saved scaler
    X_input_scaled = scaler.transform(X_input)
    X_input_tensor = torch.tensor(X_input_scaled, dtype=torch.float32).to(DEVICE)

    with torch.no_grad():
        outputs = pred_model(X_input_tensor)
        probabilities = torch.softmax(outputs, dim=1).cpu().numpy()[0]
        pred_encoded = np.argmax(probabilities)

    predicted_label = le.inverse_transform([pred_encoded])[0]
    confidence = probabilities[pred_encoded] * 100

    top3_idx = np.argsort(probabilities)[::-1][:3]
    top3_labels = le.inverse_transform(top3_idx)
    top3_probs = probabilities[top3_idx] * 100

    print("\n" + "-" * 50)
    print(f"Predicted Phase: {predicted_label}")
    print(f"Confidence: {confidence:.2f}%")
    print("\nTop-3 Candidates:")
    for lbl, prob in zip(top3_labels, top3_probs):
        bar = "█" * int(prob / 5)
        print(f"{lbl:<20} {prob:6.2f}% {bar}")
    print("-" * 50)

    again = input("\nPredict another sample? (yes/no): ").lower()
    if again not in ('yes', 'y'):
        break

print("\nExiting predictor.")