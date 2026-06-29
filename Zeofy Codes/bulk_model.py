import joblib
import numpy as np
import pandas as pd
from pathlib import Path
import time
from model_selector import ModelSelector
from main_model import (
    ZFYNeuralNet, _build_feature_vector,
    _load_version_artifacts, _infer,
    TORCH_AVAILABLE
)

try:
    import torch
except ImportError:
    pass


class BulkZeoliteModel:

    def __init__(self):
        self.model         = None
        self.scaler        = None
        self.label_encoder = None
        self.model_loaded  = False
        self._loaded_version = None
        self._is_pytorch   = False

    # --------------------------------------------------
    # LOAD MODEL FILES
    # --------------------------------------------------
    def _load_model_files(self):
        try:
            version    = ModelSelector.get_current_version()
            is_pytorch = ModelSelector.is_pytorch_model()
            label      = ModelSelector.get_display_label()
            name       = ModelSelector.get_model_name()

            print(f"\n[BulkZeoliteModel] Loading {label} ({name}) ...")
            model, scaler, label_encoder, is_pytorch = _load_version_artifacts(version)

            self.model           = model
            self.scaler          = scaler
            self.label_encoder   = label_encoder
            self._is_pytorch     = is_pytorch
            self.model_loaded    = True
            self._loaded_version = version
            print(f"  ✓ Bulk model ready: {label}")
        except Exception as e:
            self.model_loaded    = False
            self._loaded_version = None
            self._is_pytorch     = False
            print(f"\n✗ ERROR loading bulk model files: {e}\n")
            raise

    def _ensure_model(self):
        active = ModelSelector.get_current_version()
        if active != self._loaded_version:
            print(f"[BulkZeoliteModel] Switching: '{self._loaded_version}' → '{active}'")
            try:
                self._load_model_files()
            except Exception as e:
                print(f"[BulkZeoliteModel] Failed: {e}")
                return False
        return True

    def _run_inference(self, X_scaled: np.ndarray) -> np.ndarray:
        if self._is_pytorch:
            tensor = torch.tensor(X_scaled, dtype=torch.float32)
            with torch.no_grad():
                logits = self.model(tensor)
                probs  = torch.softmax(logits, dim=1).numpy()[0]
        else:
            probs = self.model.predict_proba(X_scaled)[0]
        return probs

    # --------------------------------------------------
    # PROCESS BULK FROM DATAFRAME  (single active model)
    # --------------------------------------------------
    def process_bulk_dataframe(self, df, n=5, progress_callback=None):
        if not self._ensure_model() or not self.model_loaded:
            print("ERROR: Model not loaded")
            return None
        return self._run_bulk(df, self.model, self.scaler, self.label_encoder,
                              self._is_pytorch, n=n, progress_callback=progress_callback)
    
    def process_all_models(self, df, n=5, progress_callback=None):
        """Run all 6 model versions over df and cache results.

        progress_callback(model_idx: int, label: str) is called after each
        model finishes. model_idx is 0-based so the caller can compute
        (model_idx + 1) / total to get a 0→1 fraction.
        """
        all_versions = list(ModelSelector.MODEL_CONFIGS.keys())
        all_results  = {}
        errors       = {}
        n_models     = len(all_versions)

        for i, version_key in enumerate(all_versions):
            label = ModelSelector.MODEL_CONFIGS[version_key].get("display_label", version_key)
            print(f"\n[process_all_models] ({i+1}/{n_models}) {label} ...")
            try:
                model, scaler, label_encoder, is_pytorch = _load_version_artifacts(version_key)
                result = self._run_bulk(df, model, scaler, label_encoder,
                                        is_pytorch, n=n, progress_callback=None)
                all_results[version_key] = result
                if result and result.get("success"):
                    print(f"  ✓ {label}: {result['total_observations']} obs processed")
                else:
                    print(f"  ✗ {label}: processing returned no results")
            except Exception as e:
                print(f"  ✗ {label} failed: {e}")
                import traceback; traceback.print_exc()
                errors[version_key]       = str(e)
                all_results[version_key]  = None

            # Notify caller: pass 0-based index + label of the model just finished
            if progress_callback:
                progress_callback(i, label)

        all_results["errors"] = errors
        return all_results

    # --------------------------------------------------
    # INTERNAL — run one model over the whole dataframe
    # --------------------------------------------------
    def _run_bulk(self, df, model, scaler, label_encoder, is_pytorch, n=5, progress_callback=None):
        try:
            feature_columns = ["sival", "alval", "naval", "h20", "mag", "ohval", "time", "temper"]
            total_observations = len(df)
            results            = []
            _first_logged      = [False]

            for row_num, (idx, row) in enumerate(df.iterrows()):
                if progress_callback:
                    progress = int((row_num + 1) / total_observations * 90)
                    progress_callback(progress)

                is_bad_row = bool(row.get("_row_error", False))
                params     = {col: row[col] for col in feature_columns}
                metakaolin  = row.get("metakaolin", 0.0)

                # Original raw strings (as the user typed them) — used in output Excel
                orig_params = {
                    col: row.get(f"_orig_{col}", params[col])
                    for col in feature_columns
                }
                orig_metakaolin = row.get("_orig_metakaolin", metakaolin)

                if is_bad_row:
                    # Row had null / special-char / non-numeric values — skip inference,
                    # write back exactly what the user put in the file.
                    results.append({
                        "observation_index":   row_num + 1,
                        "input_parameters":    orig_params,
                        "metakaolin":           orig_metakaolin,
                        "chemical_outputs":    {"RHA": "N/A", "AL203": "N/A",
                                                "SodiumSilicate": "N/A", "NaOH": "N/A", "Water": "N/A"},
                        "predicted_framework": "ERROR",
                        "confidence":          "N/A",
                        "top_predictions":     [],
                    })
                    continue

                chemical_outputs = self._calculate_chemical_outputs(params, metakaolin)

                molar_keys = ["sival", "alval", "naval", "mag", "h20", "ohval"]
                molar_sum  = sum(params.get(k, 0.0) for k in molar_keys)

                normalized_params = dict(params)
                if molar_sum > 0:
                    for k in molar_keys:
                        normalized_params[k] = params[k] / molar_sum
                else:
                    for k in molar_keys:
                        normalized_params[k] = 0.0

                top_predictions = self._predict_single(
                    normalized_params, model, scaler, label_encoder, is_pytorch,
                    n=n, raw_params=params
                )

                if not _first_logged[0]:
                    print(f"  Normalization check (obs 1, molar_sum={molar_sum:.4f}):")
                    for k in molar_keys:
                        print(f"    {k}: {params[k]:.4f} → {normalized_params[k]:.6f}")
                    _first_logged[0] = True

                results.append({
                    "observation_index":   row_num + 1,
                    "input_parameters":    params,
                    "metakaolin":           metakaolin,
                    "chemical_outputs":    chemical_outputs,
                    "predicted_framework": top_predictions[0][0],
                    "confidence":          top_predictions[0][1] * 100,
                    "top_predictions": [
                        {"framework": fw, "probability": conf * 100}
                        for fw, conf in top_predictions
                    ]
                })
                time.sleep(0.005)

            if progress_callback:
                progress_callback(100)

            return {
                "success":            True,
                "total_observations": total_observations,
                "results":            results
            }
        except Exception as e:
            print(f"ERROR during bulk run: {e}")
            import traceback; traceback.print_exc()
            return None

    # --------------------------------------------------
    # PREDICT SINGLE OBSERVATION
    # --------------------------------------------------
    def _predict_single(self, params, model, scaler, label_encoder, is_pytorch,
                        n=5, raw_params=None):
        rp = raw_params if raw_params is not None else params

        sival_n  = float(params.get("sival",  0.0))
        alval_n  = float(params.get("alval",  0.0))
        naval_n  = float(params.get("naval",  0.0))
        h20_n    = float(params.get("h20",    0.0))
        mag_n    = float(params.get("mag",    0.0))
        ohval_n  = float(params.get("ohval",  0.0))
        time_n   = float(params.get("time",   0.0))
        temper_n = float(params.get("temper", 0.0))

        # Ratios computed from normalized values — log1p transformed to match datasetcheck2.py
        si_al_ratio = np.log1p(sival_n / (alval_n + 1e-6))
        oh_si_ratio = np.log1p(ohval_n / (sival_n + 1e-6))
        na_al_ratio = np.log1p(naval_n / (alval_n + 1e-6))
        oh_al_ratio = np.log1p(ohval_n / (alval_n + 1e-6))

        X = np.array([[
            sival_n, alval_n, naval_n, h20_n, mag_n, ohval_n,
            time_n, temper_n,
            si_al_ratio, oh_si_ratio, na_al_ratio, oh_al_ratio
        ]], dtype=np.float32)

        X_scaled = scaler.transform(X)

        if is_pytorch:
            tensor = torch.tensor(X_scaled, dtype=torch.float32)
            with torch.no_grad():
                logits = model(tensor)
                probs  = torch.softmax(logits, dim=1).numpy()[0]
        else:
            probs = model.predict_proba(X_scaled)[0]

        top_indices    = np.argsort(probs)[-n:][::-1]
        top_frameworks = label_encoder.inverse_transform(top_indices)
        top_confs      = probs[top_indices]
        return list(zip(top_frameworks, top_confs))

    # --------------------------------------------------
    # CALCULATE CHEMICAL OUTPUTS
    # --------------------------------------------------
    def _calculate_chemical_outputs(self, params, metakaolin):
        try:
            sival = params.get("sival", 0.0)
            alval = params.get("alval", 0.0)
            naval = params.get("naval", 0.0)
            mag   = params.get("mag",   0.0)
            h20   = params.get("h20",   0.0)
            ohval = params.get("ohval", 0.0)
            gMK   = metakaolin
            # mol_total matches _build_feature_vector — all 6 chemical inputs
            mol_total = sival + alval + naval + mag + h20 + ohval
            if mol_total == 0:
                return {"RHA": 0.0, "AL203": 0.0, "SodiumSilicate": 0.0, "NaOH": 0.0, "Water": 0.0}
            # Normalized fractions — identical to _build_feature_vector
            G = sival / mol_total   # sival_n
            H = alval / mol_total   # alval_n
            I = h20   / mol_total   # h20_n
            J = naval / mol_total   # naval_n
            K = ohval / mol_total   # ohval_n
            gSS   = (J - K) / 2 * 61.98 * 100 / 10.95
            gRHA  = (G - (gMK) * (0.54 / 60.08) - (gSS) * (0.3308 / 60.08)) * (100 / 39.6) * 28.09
            gNaOH = K * 40
            gAO   = ((H / 2) - (gMK) * (0.43 / 101.96) - (gRHA) * (0.579 / 100) / 226.98 / 2) * 101.96
            gWater = (I - (gSS) * 55.97 / 100 / 18.015) * 18.015
            return {"RHA": gRHA, "AL203": gAO, "SodiumSilicate": gSS, "NaOH": gNaOH, "Water": gWater}
        except Exception as e:
            print(f"Error calculating amounts: {e}")
            return {"RHA": 0.0, "AL203": 0.0, "SodiumSilicate": 0.0, "NaOH": 0.0, "Water": 0.0}

    # --------------------------------------------------
    # SAVE RESULTS TO EXCEL  (uses one version's results)
    # --------------------------------------------------
    def save_results_to_excel(self, results, output_filename="bulk_zeolite_results.xlsx"):
        try:
            if not results or not results.get("success"):
                print("ERROR: No valid results to save")
                return False

            def _safe_round(val, digits=4):
                try:
                    return round(float(val), digits)
                except (TypeError, ValueError):
                    return val   # keep "N/A" or "ERROR" as-is

            rows = []
            for result in results["results"]:
                co = result["chemical_outputs"]
                row = {
                    "sival":          result["input_parameters"]["sival"],
                    "alval":          result["input_parameters"]["alval"],
                    "naval":          result["input_parameters"]["naval"],
                    "mag":            result["input_parameters"]["mag"],
                    "h20":            result["input_parameters"]["h20"],
                    "ohval":          result["input_parameters"]["ohval"],
                    "time":           result["input_parameters"]["time"],
                    "temper":         result["input_parameters"]["temper"],
                    "metakaolin":      result["metakaolin"],
                    "RHA":            _safe_round(co["RHA"]),
                    "AL203":          _safe_round(co["AL203"]),
                    "SodiumSilicate": _safe_round(co["SodiumSilicate"]),
                    "NaOH":           _safe_round(co["NaOH"]),
                    "Water":          _safe_round(co["Water"]),
                    "Code1":          result["predicted_framework"],
                    "ConfidenceScore": _safe_round(result["confidence"], 2),
                }
                rows.append(row)

            df_out = pd.DataFrame(rows)
            df_out.to_excel(output_filename, index=False)

            # ── Apply red font to negative values, N/A, and ERROR cells ──
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font

                RED_FONT = Font(color="FF6B6B")

                wb = load_workbook(output_filename)
                ws = wb.active

                # Column index map (1-based, skipping header row)
                col_indices = {cell.value: cell.column for cell in ws[1]}

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        val = cell.value
                        if val is None:
                            continue
                        # Flag N/A and ERROR strings
                        if isinstance(val, str) and val.strip().upper() in ("N/A", "ERROR"):
                            cell.font = RED_FONT
                        # Flag negative numbers
                        elif isinstance(val, (int, float)) and val < 0:
                            cell.font = RED_FONT

                wb.save(output_filename)
                print("  ✓ Red font applied to negative / N/A / ERROR cells")
            except Exception as style_err:
                print(f"  Warning: could not apply cell styling: {style_err}")

            print(f"Results saved to {output_filename}")
            return True
        except Exception as e:
            print(f"ERROR saving results: {e}")
            return False

    def get_framework_frequencies(self, results):
        if not results or not results.get("success"):
            return {}
        from collections import Counter
        frameworks = [
            r["predicted_framework"] for r in results["results"]
            if r["predicted_framework"] != "ERROR"
        ]
        return dict(Counter(frameworks).most_common())